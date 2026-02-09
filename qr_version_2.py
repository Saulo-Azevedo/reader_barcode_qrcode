import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import time
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


DEFAULT_DEDUP_MS = 4000


def normalize_keyboard_layout(s: str) -> str:
    """
    Normaliza bugs comuns de layout de teclado em scanner HID (PT-BR/US).
    Ex.: 'httpsÇ;;' ou 'httpsç;;' -> 'https://'
    """
    return (
        s.replace("ç", ":")
         .replace("Ç", ":")
         .replace(";", "/")
    )


class QRUsbBridgeApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Leitor QR/Barcode USB (Offline + Excel)")
        self.root.geometry("1220x700")

        # Estado
        self.running = False # aqui "rodando" significa: foco travado; export funciona sempre
        self.last_seen_at = {} # code -> monotonic ms (dedup)
        self.counts = defaultdict(int) # code -> count
        self.row_index = {} # code -> idx (ordem do primeiro aparecimento)
        self.next_idx = 1

        # Timestamps
        self.first_seen = {} # code -> datetime
        self.last_seen = {} # code -> datetime

        # Buffer oculto de captura
        self.scan_buffer = []
        self.last_key_ts = 0.0
        self.flush_after_id = None

        # Cooldown curto (reduz emenda sem ficar lento)
        self.capture_busy_until = 0.0

        # Config vars
        self.dedup_ms_var = tk.StringVar(value=str(DEFAULT_DEDUP_MS))

        # UI
        self._build_ui()

        # Foco travado somente quando rodando
        self._keep_focus()

    # ---------------- UI ----------------
    def _build_ui(self):
        outer = ttk.Frame(self.root, padding=10)
        outer.pack(fill="both", expand=True)

        top = ttk.LabelFrame(outer, text="Configuração / Operação")
        top.pack(fill="x")

        ttk.Label(top, text="DEDUP_MS:").grid(row=0, column=0, sticky="w", padx=6, pady=6)
        self.dedup_entry = ttk.Entry(top, textvariable=self.dedup_ms_var, width=10)
        self.dedup_entry.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        self.btn_start = ttk.Button(top, text="▶ Iniciar", command=self.start)
        self.btn_start.grid(row=0, column=2, padx=6, pady=6)

        self.btn_stop = ttk.Button(top, text="■ Parar", command=self.stop, state="disabled")
        self.btn_stop.grid(row=0, column=3, padx=6, pady=6)

        self.btn_export = ttk.Button(top, text="📤 Exportar Excel", command=self.export_excel)
        self.btn_export.grid(row=0, column=4, padx=6, pady=6)

        self.status_var = tk.StringVar(value="Status: parado (campos livres)")
        ttk.Label(top, textvariable=self.status_var).grid(row=0, column=5, sticky="w", padx=12, pady=6)

        scan_box = ttk.LabelFrame(outer, text="Leitura")
        scan_box.pack(fill="x", pady=(10, 0))

        ttk.Label(scan_box, text="Aponte e leia. (ENTER finaliza a leitura do scanner):").pack(anchor="w", padx=6, pady=(6, 0))

        # Campo visual fica limpo: capturamos teclado via bind e bloqueamos a inserção
        self.scan_entry = ttk.Entry(scan_box, font=("Segoe UI", 14))
        self.scan_entry.pack(fill="x", padx=6, pady=6)
        self.scan_entry.bind("<KeyPress>", self._on_keypress)

        mid = ttk.Frame(outer)
        mid.pack(fill="both", expand=True, pady=10)

        left = ttk.LabelFrame(mid, text="Tabela (leituras)")
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        cols = ("idx", "codigo", "contagem", "primeira", "ultima")
        self.tree = ttk.Treeview(left, columns=cols, show="headings", height=18)
        self.tree.heading("idx", text="#")
        self.tree.heading("codigo", text="Código (QR/Barcode)")
        self.tree.heading("contagem", text="Contagem")
        self.tree.heading("primeira", text="Primeira Leitura")
        self.tree.heading("ultima", text="Última Leitura")

        self.tree.column("idx", width=50, anchor="center")
        self.tree.column("codigo", width=520, anchor="w")
        self.tree.column("contagem", width=90, anchor="center")
        self.tree.column("primeira", width=170, anchor="center")
        self.tree.column("ultima", width=170, anchor="center")

        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        vsb.pack(side="right", fill="y", pady=6)

        btns = ttk.Frame(left)
        btns.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Button(btns, text="Limpar tabela", command=self._clear_table).pack(side="left")
        ttk.Button(btns, text="Copiar selecionado", command=self._copy_selected).pack(side="left", padx=(8, 0))

        right = ttk.LabelFrame(mid, text="Log ao vivo")
        right.pack(side="right", fill="both", expand=True)

        self.log = tk.Text(right, height=18, font=("Consolas", 10))
        self.log.pack(fill="both", expand=True, padx=6, pady=6)

    # ---------------- Focus handling ----------------
    def _keep_focus(self):
        """
        Rodando: força foco no campo de leitura.
        Parado: deixa livre (campos liberados).
        """
        try:
            if self.running:
                self.scan_entry.focus_set()
        except Exception:
            pass
        self.root.after(250, self._keep_focus)

    # ---------------- Runtime ----------------
    def start(self):
        if self.running:
            return
        self.running = True
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.status_var.set("Status: rodando (foco travado no campo de leitura)")
        self._log("✅ INICIADO. Leitura pronta (offline).")

    def stop(self):
        if not self.running:
            return
        self.running = False
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.status_var.set("Status: parado (campos livres)")
        self._log("⏹ PARADO. Você pode clicar/editar campos à vontade.")

    # ---------------- Keyboard capture ----------------
    def _on_keypress(self, event):
        # Cooldown curto para evitar emenda (não pode deixar lento)
        now = time.time()
        if now < self.capture_busy_until:
            return "break"

        key = event.keysym
        ch = event.char or ""
        self.last_key_ts = now

        # ENTER finaliza
        if key in ("Return", "KP_Enter"):
            self._finalize_scan()
            return "break"

        # Backspace
        if key == "BackSpace":
            if self.scan_buffer:
                self.scan_buffer.pop()
            return "break"

        # ignora especiais sem char
        if not ch:
            return "break"

        self.scan_buffer.append(ch)

        # Flush por silêncio (curto e responsivo)
        if self.flush_after_id is not None:
            try:
                self.root.after_cancel(self.flush_after_id)
            except Exception:
                pass

        self.flush_after_id = self.root.after(70, self._flush_if_idle)
        return "break"

    def _flush_if_idle(self):
        if time.time() - self.last_key_ts >= 0.07:
            self._finalize_scan()

    def _finalize_scan(self):
        if self.flush_after_id is not None:
            try:
                self.root.after_cancel(self.flush_after_id)
            except Exception:
                pass
            self.flush_after_id = None

        raw = "".join(self.scan_buffer).strip()
        self.scan_buffer = []

        if not raw:
            return

        # cooldown mínimo
        self.capture_busy_until = time.time() + 0.04 # 40ms

        code = normalize_keyboard_layout(raw).strip()
        if code != raw:
            self._log(f"🔧 NORMALIZADO: {raw} -> {code}")

        # Dedup
        try:
            dedup_ms = int(self.dedup_ms_var.get().strip())
        except Exception:
            dedup_ms = DEFAULT_DEDUP_MS

        now_ms = int(time.monotonic() * 1000)
        last = self.last_seen_at.get(code)
        if last is not None and (now_ms - last) < dedup_ms:
            self._log(f"🟡 DEDUP: ignorado (< {dedup_ms}ms) | {code}")
            self._bump_table(code)
            return
        self.last_seen_at[code] = now_ms

        # Marca timestamps
        now_dt = datetime.now()
        if code not in self.first_seen:
            self.first_seen[code] = now_dt
        self.last_seen[code] = now_dt

        # Sempre registra na tabela (offline)
        self._log(f"📥 CAPTURADO: {code}")
        self._bump_table(code)

    # ---------------- Table helpers ----------------
    def _bump_table(self, code: str):
        self.counts[code] += 1
        count = self.counts[code]

        if code not in self.row_index:
            self.row_index[code] = self.next_idx
            self.next_idx += 1

        idx = self.row_index[code]

        primeira = self.first_seen.get(code)
        ultima = self.last_seen.get(code)

        primeira_s = primeira.strftime("%d/%m/%Y %H:%M:%S") if primeira else "-"
        ultima_s = ultima.strftime("%d/%m/%Y %H:%M:%S") if ultima else "-"

        iid = self._find_iid_by_code(code)
        if iid is None:
            self.tree.insert("", "end", values=(idx, code, count, primeira_s, ultima_s))
        else:
            self.tree.item(iid, values=(idx, code, count, primeira_s, ultima_s))

    def _find_iid_by_code(self, code: str):
        for iid in self.tree.get_children(""):
            vals = self.tree.item(iid, "values")
            if vals and vals[1] == code:
                return iid
        return None

    def _clear_table(self):
        for iid in self.tree.get_children(""):
            self.tree.delete(iid)
        self.counts.clear()
        self.row_index.clear()
        self.last_seen_at.clear()
        self.first_seen.clear()
        self.last_seen.clear()
        self.next_idx = 1
        self._log("🧹 Tabela/contadores/timestamps limpos.")

    def _copy_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        code = vals[1]
        self.root.clipboard_clear()
        self.root.clipboard_append(code)
        self._log(f"📋 Copiado: {code}")

    # ---------------- Export Excel ----------------
    def export_excel(self):
        if not self.counts:
            messagebox.showwarning("Exportar Excel", "Nenhuma leitura para exportar.")
            return

        default_name = f"leituras_qr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel (*.xlsx)", "*.xlsx")],
            title="Salvar relatório de leituras"
        )
        if not filepath:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Leituras"

        # Cabeçalho
        headers = ["#", "Código", "Quantidade", "Primeira Leitura", "Última Leitura"]
        ws.append(headers)

        # Linhas (ordenadas pelo índice)
        for code, idx in sorted(self.row_index.items(), key=lambda x: x[1]):
            primeira = self.first_seen.get(code)
            ultima = self.last_seen.get(code)

            ws.append([
                idx,
                code,
                self.counts.get(code, 0),
                primeira.strftime("%d/%m/%Y %H:%M:%S") if primeira else "",
                ultima.strftime("%d/%m/%Y %H:%M:%S") if ultima else "",
            ])

        # Ajuste de largura automático simples
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[letter].width = min(max_len + 2, 80)

        try:
            wb.save(filepath)
            self._log(f"✅ Excel exportado: {filepath}")
            messagebox.showinfo("Exportar Excel", f"Arquivo salvo com sucesso:\n{filepath}")
        except Exception as e:
            self._log(f"❌ Erro ao salvar Excel: {e}")
            messagebox.showerror("Erro ao salvar", str(e))

    # ---------------- Logging ----------------
    def _log(self, msg: str):
        ts = time.strftime("%H:%M:%S")
        self.log.insert("end", f"[{ts}] {msg}\n")
        self.log.see("end")


def main():
    root = tk.Tk()
    try:
        ttk.Style().theme_use("clam")
    except Exception:
        pass

    app = QRUsbBridgeApp(root)
    try:
        root.mainloop()
    except KeyboardInterrupt:
        # evita stacktrace se parar no console
        pass


if __name__ == "__main__":
    main()